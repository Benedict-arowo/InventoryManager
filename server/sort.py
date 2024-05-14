from dataclasses import dataclass
import datetime
import os
import pprint
from tracemalloc import start
from django.utils import timezone
from openpyxl import Workbook

from openpyxl.styles import Font, Alignment
from openpyxl.styles import NamedStyle

ACCOUNTER_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "accounter.settings")

import django

django.setup()

from main.models import Sale, Stock, User

separtor = "*~*"
title_font_size = 16

main_row = 1
main_col = 1


def arrange_data(data):
    arranged_data = {}
    for i in data:
        item_name = i.item.name

        # If the item has no category, then it is uncategorized
        if i.item.category == None:
            item_category = "Uncategorized"
        else:
            item_category = i.item.category.name.capitalize()

        if item_category in arranged_data:
            # Add the item to the existing set
            arranged_data[item_category].add(item_name)
        else:
            # Initialize a new set, and put the item in the new set
            arranged_data[item_category] = {item_name}

    for row in arranged_data:
        arranged_data[row] = list(arranged_data[row])

    # pprint.pp(arranged_data)
    return arranged_data


def get_data(start_day, end_day, month, year, type="arranged"):
    # TODO: support for either Sales or Expences
    # if not end_date_str:
    #     filters["created_at__date"] = date
    # if end_date_str:
    # if name:
    # filters["name__icontains"] = name
    filters = {}

    start_date = datetime.datetime.strptime(
        year + "-" + month + "-" + start_day, "%Y-%m-%d"
    ).date()
    end_date = datetime.datetime.strptime(
        year + "-" + month + "-" + end_day, "%Y-%m-%d"
    ).date()

    print(start_date, end_date)
    filters["created_at__date__range"] = [start_date, end_date]

    # Filter sales based on the provided date
    data = Sale.objects.filter(**filters).order_by("created_at")

    return dict(arranged=arrange_data(data), raw=data)


def arrange_data_by_date(data):
    arranged_data = {}
    for i in data:
        item_date = i.created_at.date()
        item_name = i.item.name
        # print(datetime.datetime(2024, 3, 1, 0, 0, tzinfo=datetime.timezone.utc).date())
        if item_date in arranged_data:
            # Add the item to the existing set
            if item_name in arranged_data[item_date]:
                arranged_data[item_date][item_name] = {
                    "quantity": arranged_data[item_date][item_name]["quantity"]
                    + i.quantity,
                    "total": arranged_data[item_date][item_name]["total"] + i.total,
                }
            else:
                arranged_data[item_date][item_name] = {
                    "quantity": i.quantity,
                    "total": i.total,
                }
        else:
            # Initialize a new array, and put the item in the new array
            arranged_data[item_date] = {
                item_name: {
                    "quantity": i.quantity,
                    "total": i.total,
                }
            }
    return arranged_data


def convert_to_excel(arranged_data, arranged_by_date):

    wb = Workbook()

    ws = wb.active
    ws.column_dimensions["A"].width = 25
    indexes = {}

    while True:
        row_count = main_row
        col_count = main_col
        for row in arranged_data:
            date_style = NamedStyle(name="date_style")
            date_style.number_format = "YYYY-MM-DD"  # Set the date format

            cell = ws.cell(row=row_count, column=col_count, value=row)
            cell.font = Font(bold=True, size=title_font_size)
            # cell.style = date_style
            # cell.alignment = Alignment(horizontal="center")
            row_count += 1
            # col_count += 1

            for child_row in arranged_data[row]:
                cell = ws.cell(
                    row=row_count, column=col_count, value=child_row.capitalize()
                )
                indexes[child_row] = [col_count, row_count]
                row_count += 1
            # col_count -= 1
        # pprint.pp(indexes)
        break

    while True:
        row_count = main_row
        col_count = main_col + 1
        for row in arranged_by_date:
            cell = ws.cell(row=row_count, column=col_count, value=row)
            cell.font = Font(bold=True, size=12)

            for child_row in arranged_by_date[row]:
                item = arranged_by_date[row][child_row]
                row_count += 1
                if child_row in indexes:
                    index_row_count = indexes[child_row][1]
                    cell = ws.cell(
                        row=index_row_count, column=col_count, value=item["total"]
                    )
                else:
                    print("Error adding an item -", child_row)

            col_count += 1
            row_count = 1
        break

    wb.save("TestData.xlsx")


if __name__ == "__main__":
    # start_day = int(input("Start Day - "))
    # end_day = int(input("End Day - "))
    # month = int(input("Month - "))
    # year = int(input("Year - "))
    # type = int(input("Type - "))

    start_day = "01"
    end_day = "29"
    month = "02"
    year = "2024"
    # type = "sales"

    data = get_data(start_day, end_day, month, year)
    arranged_by_date = arrange_data_by_date(data["raw"])
    convert_to_excel(data["arranged"], arranged_by_date)
