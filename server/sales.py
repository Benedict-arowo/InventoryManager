from datetime import datetime
import os
from django.utils import timezone
import openpyxl

ACCOUNTER_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "accounter.settings")

import django

django.setup()

from main.models import Sale, Stock, User


def check_equation(row):
    numerator, denominator = row.replace("=", "").split("/")
    return float(numerator) / float(denominator)


def create_items(filepath, year, month):
    wb_obj = openpyxl.load_workbook(filepath)
    sheets = wb_obj.sheetnames[1:]
    itemList = []

    for sheet in sheets:
        activeSheet = wb_obj[sheet]
        wb_obj.active = activeSheet

        sheet_obj = wb_obj.active

        row = sheet_obj.max_row

        cell_obj = sheet_obj["A2:D{}".format(row)]

        total = 0
        sheet_name = sheet_obj.title.split("-")
        day = sheet_name[1]
        if day == "#":
            break
        record_date = timezone.make_aware(datetime(year, month, int(day), 0, 0, 0))

        for row in cell_obj:

            start_datetime = timezone.make_aware(
                timezone.datetime.combine(record_date, timezone.datetime.min.time())
            )
            end_datetime = timezone.make_aware(
                timezone.datetime.combine(record_date, timezone.datetime.max.time())
            )
            Sale.objects.filter(
                created_at__range=(start_datetime, end_datetime)
            ).delete()

            id, item_name, sales_price, quantity_sold = row
            if quantity_sold.value != None:
                if (
                    not isinstance(sales_price.value, int)
                    and not isinstance(sales_price.value, float)
                    and "=" in sales_price.value
                ):
                    numerator, denominator = sales_price.value.replace("=", "").split(
                        "/"
                    )
                    sales_price.value = check_equation(sales_price.value)

                if (
                    not isinstance(quantity_sold.value, int)
                    and not isinstance(quantity_sold.value, float)
                    and "=" in quantity_sold.value
                ):
                    quantity_sold.value = check_equation(quantity_sold.value)

                item, _ = Stock.objects.get_or_create(name=item_name.value.lower())
                itemList.append(
                    Sale(
                        item=item,
                        quantity=quantity_sold.value,
                        amount_paid=quantity_sold.value * sales_price.value,
                        status="Paid",
                        price=sales_price.value,
                        total=quantity_sold.value * sales_price.value,
                        sold_by=User.objects.get(username="admin"),
                        created_at=record_date,
                    )
                )

                total += quantity_sold.value * sales_price.value
        print(f"{activeSheet.title} - {total}")

    if len(itemList) != 0:
        Sale._meta.get_field("created_at").auto_now_add = False
        Sale.objects.bulk_create(itemList)
        Sale._meta.get_field("created_at").auto_now_add = True


if __name__ == "__main__":
    folder_name = input("Folder name? ")
    rootDir = os.path.dirname(os.path.abspath(__file__))
    parentDir = os.path.abspath(os.path.join(rootDir, os.pardir))
    filesDir = os.path.join(parentDir, "csv_files", "Sales", folder_name)
    year = int(input("Enter year for file: "))
    month = int(input("Enter month for file: "))

    for subdir, dirs, files in os.walk(filesDir):
        for file in files:
            print(os.path.join(filesDir, file))
            create_items(os.path.join(filesDir, file), year, month)
