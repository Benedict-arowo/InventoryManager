from datetime import datetime
import os
import openpyxl
from django.utils import timezone


ACCOUNTER_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "accounter.settings")

import django

django.setup()

from main.models import Sale, Stock, User, Expence


def check_equation(row):
    if "/" in row:
        numerator, denominator = row.replace("=", "").split("/")
        return float(numerator) / float(denominator)
    elif "*" in row:
        val1, val2 = row.replace("=", "").split("*")
        return float(val1) * float(val2)
    elif "+" in row:
        val1, val2 = row.replace("=", "").split("+")
        return float(val1) + float(val2)
    elif "-" in row:
        val1, val2 = row.replace("=", "").split("-")
        return float(val1) - float(val2)
    else:
        return None


def create_items(filepath, year, month):
    wb_obj = openpyxl.load_workbook(filepath)
    sheets = wb_obj.sheetnames
    item_list = []

    for sheet in sheets:
        if sheet.lower() != "summary":
            wb_obj.active = wb_obj[sheet]
            active_sheet = wb_obj.active
            cell_obj = active_sheet["A2:D{}".format(active_sheet.max_row)]

            total = 0

            day = sheet.split("-")[1]
            if day == "#":
                break
            record_date = timezone.make_aware(datetime(year, month, int(day), 0, 0, 0))

            for row in cell_obj:
                start_datetime = timezone.make_aware(
                    timezone.datetime.combine(record_date, timezone.datetime.min.time())
                )
                end_datetime = timezone.make_aware(
                    timezone.datetime.combine(record_date, timezone.datetime.max.time())
                )
                Expence.objects.filter(
                    created_at__range=(start_datetime, end_datetime)
                ).delete()

                id, item_name, price_per_unit, quantity_bought = row
                if quantity_bought.value != None:
                    if (
                        not isinstance(price_per_unit.value, int)
                        and not isinstance(price_per_unit.value, float)
                        and "=" in price_per_unit.value
                    ):
                        price_per_unit.value = check_equation(price_per_unit.value)

                    if (
                        not isinstance(quantity_bought.value, int)
                        and not isinstance(quantity_bought.value, float)
                        and "=" in quantity_bought.value
                    ):
                        quantity_bought.value = check_equation(quantity_bought.value)

                    item, _ = Stock.objects.get_or_create(name=item_name.value.lower())
                    item_list.append(
                        Expence(
                            item=item,
                            quantity=quantity_bought.value,
                            price_per_unit=price_per_unit.value,
                            total=quantity_bought.value * price_per_unit.value,
                            created_at=record_date,
                        )
                    )
                    total += quantity_bought.value * price_per_unit.value
            print(f"{sheet} - {total}")

    #
    if len(item_list) != 0:
        Expence._meta.get_field("created_at").auto_now_add = False
        Expence.objects.bulk_create(item_list)
        Expence._meta.get_field("created_at").auto_now_add = True


if __name__ == "__main__":
    folder_name = input("Folder name? ")
    rootDir = os.path.dirname(os.path.abspath(__file__))
    parentDir = os.path.abspath(os.path.join(rootDir, os.pardir))
    filesDir = os.path.join(parentDir, "csv_files", "Purchases", folder_name)
    useFoldername = input("Get date from foldername MM-YY? (y/n) ")
    if not useFoldername:
        year = int(input("Enter year for file: "))
        month = int(input("Enter month for file: "))

    for subdir, dirs, files in os.walk(filesDir):
        for file in files:
            print(os.path.join(filesDir, file))
            if useFoldername:
                month, year = folder_name.split("-")
                year = int(year)
                month = int(month)
            print(file, year, month)
            create_items(os.path.join(filesDir, file), year, month)
